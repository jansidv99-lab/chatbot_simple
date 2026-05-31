import datetime
import io
import os
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from ingestion.parser import parse_positions_excel, validate_file

# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------

_XLSX_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "raw_data_files",
    "daily_poistions",
    "positions.xlsx",
)

_PNL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "raw_data_files",
    "daily_pl",
    "pnl.xlsx",
)

_TRADEBOOK_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "raw_data_files",
    "trade_book",
    "tradebook.xlsx",
)


@pytest.fixture(scope="module")
def real_file_bytes() -> bytes:
    with open(_XLSX_PATH, "rb") as f:
        return f.read()


@pytest.fixture(scope="module")
def pnl_file_bytes() -> bytes:
    with open(_PNL_PATH, "rb") as f:
        return f.read()


@pytest.fixture(scope="module")
def tradebook_file_bytes() -> bytes:
    with open(_TRADEBOOK_PATH, "rb") as f:
        return f.read()


def _make_minimal_xlsx(include_fo_sheet: bool = True, add_data_row: bool = True) -> bytes:
    wb = openpyxl.Workbook()
    if include_fo_sheet:
        ws = wb.active
        ws.title = "F&O"
        ws["B15"] = "Symbol"
        if add_data_row:
            ws["B16"] = "TESTSYM"
            ws["C16"] = "2026-01-01"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_minimal_pnl_xlsx(include_fo_sheet: bool = True, with_title: bool = True) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    if include_fo_sheet:
        ws.title = "F&O"
        if with_title:
            ws["B11"] = "P&L Statement for F&O from 2026-01-01 to 2026-01-01"
        ws["B38"] = "Symbol"
        ws["B39"] = "TESTSYM"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_minimal_tradebook_xlsx(include_fo_sheet: bool = True) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    if include_fo_sheet:
        ws.title = "F&O"
        ws["B15"] = "Symbol"
        ws["B16"] = "TESTSYM"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Positions parser tests — against the real Excel file
# ---------------------------------------------------------------------------


def test_parse_returns_correct_row_count(real_file_bytes):
    rows = parse_positions_excel(real_file_bytes)
    assert len(rows) == 5


def test_parse_maps_symbol_correctly(real_file_bytes):
    rows = parse_positions_excel(real_file_bytes)
    assert rows[0]["symbol"] == "NIFTY26MAY23050PE"


def test_parse_maps_trade_date_as_date_obj(real_file_bytes):
    rows = parse_positions_excel(real_file_bytes)
    assert isinstance(rows[0]["trade_date"], datetime.date)
    assert rows[0]["trade_date"] == datetime.date(2026, 5, 19)


def test_parse_excludes_client_id(real_file_bytes):
    rows = parse_positions_excel(real_file_bytes)
    for row in rows:
        assert "client_id" not in row


def test_parse_numeric_fields_are_float(real_file_bytes):
    rows = parse_positions_excel(real_file_bytes)
    numeric_keys = [
        "open_quantity", "open_average", "open_value",
        "previous_close_price", "closing_value",
        "unrealized_profit", "unrealized_profit_pct",
    ]
    for key in numeric_keys:
        assert isinstance(rows[0][key], float), f"{key} is not float"


def test_to_date_handles_excel_serial_number():
    from ingestion.parser import _to_date
    # Excel serial 45761 = 2025-03-01 (days since 1899-12-30)
    result = _to_date(45761)
    assert isinstance(result, datetime.date)
    assert result == datetime.date(1899, 12, 30) + datetime.timedelta(days=45761)


# ---------------------------------------------------------------------------
# Positions validation tests — using synthetic workbooks
# ---------------------------------------------------------------------------


def test_validate_rejects_non_xlsx():
    with pytest.raises(ValueError, match=r"\.xlsx"):
        validate_file("positions.csv", b"fake")


def test_validate_rejects_missing_fo_sheet():
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="F&O"):
        validate_file("test.xlsx", buf.getvalue())


def test_validate_rejects_empty_data():
    xlsx = _make_minimal_xlsx(include_fo_sheet=True, add_data_row=False)
    with pytest.raises(ValueError, match="No data rows"):
        validate_file("test.xlsx", xlsx)


# ---------------------------------------------------------------------------
# P&L parser tests — against the real pnl.xlsx
# ---------------------------------------------------------------------------


def test_parse_pnl_returns_6_pl_rows(pnl_file_bytes):
    from ingestion.parser import parse_pnl_excel
    pl_rows, _ = parse_pnl_excel(pnl_file_bytes)
    assert len(pl_rows) == 6


def test_parse_pnl_trade_date_is_date_obj(pnl_file_bytes):
    from ingestion.parser import parse_pnl_excel
    pl_rows, charges = parse_pnl_excel(pnl_file_bytes)
    assert isinstance(pl_rows[0]["trade_date"], datetime.date)
    assert pl_rows[0]["trade_date"] == datetime.date(2026, 5, 19)
    assert charges["date"] == datetime.date(2026, 5, 19)


def test_parse_pnl_charges_has_all_keys(pnl_file_bytes):
    from ingestion.parser import parse_pnl_excel, _CHARGE_KEYS
    _, charges = parse_pnl_excel(pnl_file_bytes)
    for key in _CHARGE_KEYS:
        assert key in charges, f"Missing charge key: {key}"
    assert "date" in charges


def test_parse_pnl_charges_brokerage_value(pnl_file_bytes):
    from ingestion.parser import parse_pnl_excel
    _, charges = parse_pnl_excel(pnl_file_bytes)
    assert charges["brokerage"] == 400.0


def test_parse_pnl_excludes_client_id(pnl_file_bytes):
    from ingestion.parser import parse_pnl_excel
    pl_rows, charges = parse_pnl_excel(pnl_file_bytes)
    for row in pl_rows:
        assert "client_id" not in row
    assert "client_id" not in charges


# ---------------------------------------------------------------------------
# P&L validation tests — using synthetic workbooks
# ---------------------------------------------------------------------------


def test_validate_pnl_rejects_non_xlsx():
    from ingestion.parser import validate_pnl_file
    with pytest.raises(ValueError, match=r"\.xlsx"):
        validate_pnl_file("pnl.csv", b"fake")


def test_validate_pnl_rejects_missing_fo_sheet():
    from ingestion.parser import validate_pnl_file
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="F&O"):
        validate_pnl_file("test.xlsx", buf.getvalue())


def test_validate_pnl_rejects_wrong_header():
    from ingestion.parser import validate_pnl_file
    xlsx = _make_minimal_pnl_xlsx(include_fo_sheet=True, with_title=False)
    with pytest.raises(ValueError, match="P&L Statement"):
        validate_pnl_file("test.xlsx", xlsx)


# ---------------------------------------------------------------------------
# Tradebook parser tests — against the real tradebook.xlsx
# ---------------------------------------------------------------------------


def test_parse_tradebook_aggregates_to_fewer_rows(tradebook_file_bytes):
    from ingestion.parser import parse_tradebook_excel
    rows = parse_tradebook_excel(tradebook_file_bytes)
    # 105 raw trade rows → 29 aggregated (symbol+date combos)
    assert len(rows) == 29


def test_parse_tradebook_quantity_is_summed(tradebook_file_bytes):
    from ingestion.parser import parse_tradebook_excel
    rows = parse_tradebook_excel(tradebook_file_bytes)
    by_sym = {r["symbol"]: r for r in rows}
    # NIFTY2651923600CE 2026-05-19 has 9 raw trades summing to qty=910
    assert by_sym["NIFTY2651923600CE"]["quantity"] == 910.0


def test_parse_tradebook_price_is_averaged(tradebook_file_bytes):
    from ingestion.parser import parse_tradebook_excel
    rows = parse_tradebook_excel(tradebook_file_bytes)
    by_sym = {r["symbol"]: r for r in rows}
    # NIFTY2651923600CE avg price across 9 trades rounds to 180.9222
    assert abs(by_sym["NIFTY2651923600CE"]["price"] - 180.9222) < 0.001


def test_parse_tradebook_execution_time_is_max(tradebook_file_bytes):
    from ingestion.parser import parse_tradebook_excel
    rows = parse_tradebook_excel(tradebook_file_bytes)
    by_sym = {r["symbol"]: r for r in rows}
    # NIFTY2651923600CE max order_execution_time = 2026-05-19T10:57:48
    assert by_sym["NIFTY2651923600CE"]["order_execution_time"] == datetime.datetime(2026, 5, 19, 10, 57, 48)


# ---------------------------------------------------------------------------
# Tradebook validation tests — using synthetic workbooks
# ---------------------------------------------------------------------------


def test_validate_tradebook_rejects_non_xlsx():
    from ingestion.parser import validate_tradebook_file
    with pytest.raises(ValueError, match=r"\.xlsx"):
        validate_tradebook_file("trades.csv", b"fake")


def test_validate_tradebook_rejects_missing_fo_sheet():
    from ingestion.parser import validate_tradebook_file
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="F&O"):
        validate_tradebook_file("test.xlsx", buf.getvalue())


# ---------------------------------------------------------------------------
# DB layer tests — fully mocked
# ---------------------------------------------------------------------------


def _make_mock_conn(returned_rows=None):
    mock_conn = MagicMock()
    mock_cursor = MagicMock()
    mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cursor)
    mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
    return mock_conn, mock_cursor


def test_ensure_schema_creates_all_four_tables():
    from ingestion.db import ensure_schema

    mock_conn, mock_cursor = _make_mock_conn()
    ensure_schema(mock_conn)

    all_sql = " ".join(call[0][0] for call in mock_cursor.execute.call_args_list)
    assert "daily_positions" in all_sql
    assert "daily_pl" in all_sql
    assert "daily_trades" in all_sql
    assert "daily_charges" in all_sql
    mock_conn.commit.assert_called_once()


def test_insert_returns_inserted_count():
    from ingestion.db import insert_positions

    mock_conn, _ = _make_mock_conn()
    rows = [{"symbol": "A", "trade_date": datetime.date(2026, 1, 1),
             "segment": None, "position_type": None, "open_quantity": None,
             "open_average": None, "open_value": None, "previous_close_price": None,
             "closing_value": None, "unrealized_profit": None, "unrealized_profit_pct": None}
            for _ in range(3)]

    # execute_values returns the RETURNING rows; 3 returned = 3 inserted
    with patch("ingestion.db.execute_values", return_value=[("A",), ("A",), ("A",)]):
        inserted, skipped = insert_positions(mock_conn, rows)

    assert inserted == 3
    assert skipped == 0


def test_insert_returns_skipped_count():
    from ingestion.db import insert_positions

    mock_conn, _ = _make_mock_conn()
    rows = [{"symbol": "A", "trade_date": datetime.date(2026, 1, 1),
             "segment": None, "position_type": None, "open_quantity": None,
             "open_average": None, "open_value": None, "previous_close_price": None,
             "closing_value": None, "unrealized_profit": None, "unrealized_profit_pct": None}
            for _ in range(2)]

    # execute_values returns empty list when all rows conflict
    with patch("ingestion.db.execute_values", return_value=[]):
        inserted, skipped = insert_positions(mock_conn, rows)

    assert inserted == 0
    assert skipped == 2


def test_insert_commits_on_success():
    from ingestion.db import insert_positions

    mock_conn, _ = _make_mock_conn()
    rows = [{"symbol": "X", "trade_date": datetime.date(2026, 1, 1),
             "segment": None, "position_type": None, "open_quantity": None,
             "open_average": None, "open_value": None, "previous_close_price": None,
             "closing_value": None, "unrealized_profit": None, "unrealized_profit_pct": None}]

    with patch("ingestion.db.execute_values", return_value=[("X",)]):
        insert_positions(mock_conn, rows)

    mock_conn.commit.assert_called_once()


def test_insert_pl_returns_inserted_and_skipped():
    from ingestion.db import insert_pl

    mock_conn, _ = _make_mock_conn()
    rows = [{"symbol": "A", "trade_date": datetime.date(2026, 1, 1),
             "quantity": None, "buy_value": None, "sell_value": None,
             "realized_pnl": None, "realized_pnl_pct": None, "previous_closing_price": None,
             "open_quantity": None, "open_quantity_type": None, "open_value": None,
             "unrealized_pnl": None, "unrealized_pnl_pct": None}]

    with patch("ingestion.db.execute_values", return_value=[("A",)]):
        inserted, skipped = insert_pl(mock_conn, rows)

    assert inserted == 1
    assert skipped == 0


def test_insert_trades_returns_inserted_and_skipped():
    from ingestion.db import insert_trades

    mock_conn, _ = _make_mock_conn()
    rows = [{"symbol": "A", "trade_date": datetime.date(2026, 1, 1),
             "trade_type": "buy", "quantity": 100.0, "price": 50.0,
             "order_execution_time": datetime.datetime(2026, 1, 1, 9, 0, 0)}]

    with patch("ingestion.db.execute_values", return_value=[]):
        inserted, skipped = insert_trades(mock_conn, rows)

    assert inserted == 0
    assert skipped == 1


def test_insert_charges_returns_1_0_on_new_date():
    from ingestion.db import insert_charges

    mock_conn, mock_cursor = _make_mock_conn()
    mock_cursor.fetchall.return_value = [(datetime.date(2026, 1, 1),)]
    charges = {
        "date": datetime.date(2026, 1, 1),
        "brokerage": 400.0, "exchange_transaction_charges": 276.97,
        "clearing_charges": 0.0, "central_gst": 0.0, "state_gst": 0.0,
        "integrated_gst": 122.0, "securities_transaction_tax": 564.0,
        "sebi_turnover_fees": 0.78, "stamp_duty": 12.0, "ipft": 0.0008,
    }
    inserted, skipped = insert_charges(mock_conn, charges)
    assert inserted == 1
    assert skipped == 0
