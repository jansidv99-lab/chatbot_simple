import datetime
import re
from io import BytesIO

import openpyxl

# ── Positions ────────────────────────────────────────────────────────────────

_SHEET = "F&O"
_HEADER_ROW = 15
_DATA_START_ROW = 16
_COL_SYMBOL = 2  # B
_COL_LAST = 12  # L


def validate_file(filename: str, file_bytes: bytes) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError(f"Expected an .xlsx file, got '{filename}'.")
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    if _SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{_SHEET}' not found in workbook (found: {wb.sheetnames}).")
    ws = wb[_SHEET]
    header_b = ws.cell(row=_HEADER_ROW, column=_COL_SYMBOL).value
    if str(header_b or "").strip().lower() != "symbol":
        raise ValueError(
            f"Expected 'Symbol' in cell B{_HEADER_ROW}, got '{header_b}'. "
            "File layout may not match the expected format."
        )
    has_data = any(
        ws.cell(row=r, column=_COL_SYMBOL).value not in (None, "")
        for r in range(_DATA_START_ROW, _DATA_START_ROW + 5)
    )
    if not has_data:
        raise ValueError("No data rows found after the header row.")


def _to_date(value) -> datetime.date:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date: integer days since 1899-12-30
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value))
    return datetime.date.fromisoformat(str(value).strip()[:10])


def parse_positions_excel(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[_SHEET]
    rows = []
    for row in ws.iter_rows(min_row=_DATA_START_ROW, min_col=_COL_SYMBOL, max_col=_COL_LAST):
        symbol = row[0].value
        if symbol is None or str(symbol).strip() == "":
            continue
        rows.append(
            {
                "symbol": str(symbol).strip(),
                "trade_date": _to_date(row[1].value),
                "segment": str(row[2].value).strip() if row[2].value is not None else None,
                "position_type": str(row[3].value).strip() if row[3].value is not None else None,
                "open_quantity": float(row[4].value) if row[4].value is not None else None,
                "open_average": float(row[5].value) if row[5].value is not None else None,
                "open_value": float(row[6].value) if row[6].value is not None else None,
                "previous_close_price": float(row[7].value) if row[7].value is not None else None,
                "closing_value": float(row[8].value) if row[8].value is not None else None,
                "unrealized_profit": float(row[9].value) if row[9].value is not None else None,
                "unrealized_profit_pct": float(row[10].value) if row[10].value is not None else None,
            }
        )
    return rows


# ── P&L ─────────────────────────────────────────────────────────────────────

_PNL_SHEET = "F&O"
_PNL_TITLE_ROW = 11
_CHARGES_ROW_START = 24
_PNL_DATA_START = 39
_PNL_COL_START = 2   # B
_PNL_COL_END = 14    # N

_CHARGE_KEYS = [
    "brokerage",
    "exchange_transaction_charges",
    "clearing_charges",
    "central_gst",
    "state_gst",
    "integrated_gst",
    "securities_transaction_tax",
    "sebi_turnover_fees",
    "stamp_duty",
    "ipft",
]


def validate_pnl_file(filename: str, file_bytes: bytes) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError(f"Expected an .xlsx file, got '{filename}'.")
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    if _PNL_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{_PNL_SHEET}' not found in workbook (found: {wb.sheetnames}).")
    ws = wb[_PNL_SHEET]
    title = ws.cell(row=_PNL_TITLE_ROW, column=2).value
    if title is None or "P&L Statement" not in str(title):
        raise ValueError(
            f"Expected 'P&L Statement' in cell B{_PNL_TITLE_ROW}, got '{title}'. "
            "File layout may not match the expected format."
        )
    header_b = ws.cell(row=38, column=2).value
    if str(header_b or "").strip().lower() != "symbol":
        raise ValueError(
            f"Expected 'Symbol' in cell B38, got '{header_b}'. "
            "File layout may not match the expected format."
        )


def parse_pnl_excel(file_bytes: bytes) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[_PNL_SHEET]

    title = str(ws.cell(row=_PNL_TITLE_ROW, column=2).value or "")
    dates_found = re.findall(r"\d{4}-\d{2}-\d{2}", title)
    trade_date = datetime.date.fromisoformat(dates_found[0])

    # Fixed-position charges rows 24–33, amount in col C
    charge_values = [ws.cell(row=r, column=3).value for r in range(_CHARGES_ROW_START, _CHARGES_ROW_START + 10)]
    charges: dict = {
        key: (float(v) if v is not None else None)
        for key, v in zip(_CHARGE_KEYS, charge_values)
    }
    charges["date"] = trade_date

    # P&L rows starting at row 39, cols B–N
    pl_rows = []
    for row in ws.iter_rows(min_row=_PNL_DATA_START, min_col=_PNL_COL_START, max_col=_PNL_COL_END):
        symbol = row[0].value
        if symbol is None or str(symbol).strip() == "":
            continue
        pl_rows.append(
            {
                "symbol": str(symbol).strip(),
                "trade_date": trade_date,
                # idx 1 = ISIN — skipped
                "quantity": float(row[2].value) if row[2].value is not None else None,
                "buy_value": float(row[3].value) if row[3].value is not None else None,
                "sell_value": float(row[4].value) if row[4].value is not None else None,
                "realized_pnl": float(row[5].value) if row[5].value is not None else None,
                "realized_pnl_pct": float(row[6].value) if row[6].value is not None else None,
                "previous_closing_price": float(row[7].value) if row[7].value is not None else None,
                "open_quantity": float(row[8].value) if row[8].value is not None else None,
                "open_quantity_type": str(row[9].value).strip() if row[9].value is not None else None,
                "open_value": float(row[10].value) if row[10].value is not None else None,
                "unrealized_pnl": float(row[11].value) if row[11].value is not None else None,
                "unrealized_pnl_pct": float(row[12].value) if row[12].value is not None else None,
            }
        )
    return pl_rows, charges


# ── Trade Book ───────────────────────────────────────────────────────────────

_TB_SHEET = "F&O"
_TB_HEADER_ROW = 15
_TB_DATA_START = 16
_TB_COL_START = 2   # B
_TB_COL_END = 14    # N


def validate_tradebook_file(filename: str, file_bytes: bytes) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError(f"Expected an .xlsx file, got '{filename}'.")
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    if _TB_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{_TB_SHEET}' not found in workbook (found: {wb.sheetnames}).")
    ws = wb[_TB_SHEET]
    header_b = ws.cell(row=_TB_HEADER_ROW, column=_TB_COL_START).value
    if str(header_b or "").strip().lower() != "symbol":
        raise ValueError(
            f"Expected 'Symbol' in cell B{_TB_HEADER_ROW}, got '{header_b}'. "
            "File layout may not match the expected format."
        )


def _to_datetime(value) -> datetime.datetime | None:
    if isinstance(value, datetime.datetime):
        return value
    if value is None:
        return None
    return datetime.datetime.fromisoformat(str(value).strip())


def parse_tradebook_excel(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[_TB_SHEET]

    # Accumulate per (symbol, trade_date)
    # Row layout (0-indexed from col B):
    #   0=Symbol, 1=ISIN, 2=Trade Date, 3=Exchange, 4=Segment, 5=Series,
    #   6=Trade Type, 7=Auction, 8=Quantity, 9=Price, 10=Trade ID,
    #   11=Order ID, 12=Order Execution Time
    agg: dict[tuple, dict] = {}

    for row in ws.iter_rows(min_row=_TB_DATA_START, min_col=_TB_COL_START, max_col=_TB_COL_END):
        symbol = row[0].value
        if symbol is None or str(symbol).strip() == "":
            continue
        symbol = str(symbol).strip()
        trade_date = _to_date(row[2].value)
        key = (symbol, trade_date)

        trade_type = str(row[6].value).strip() if row[6].value is not None else None
        quantity = float(row[8].value) if row[8].value is not None else 0.0
        price = float(row[9].value) if row[9].value is not None else 0.0
        exec_time = _to_datetime(row[12].value)

        if key not in agg:
            agg[key] = {
                "symbol": symbol,
                "trade_date": trade_date,
                "trade_type": trade_type,
                "_qty": quantity,
                "_price_sum": price,
                "_price_count": 1,
                "order_execution_time": exec_time,
            }
        else:
            agg[key]["_qty"] += quantity
            agg[key]["_price_sum"] += price
            agg[key]["_price_count"] += 1
            if exec_time is not None:
                current = agg[key]["order_execution_time"]
                if current is None or exec_time > current:
                    agg[key]["order_execution_time"] = exec_time

    result = []
    for state in agg.values():
        count = state.pop("_price_count")
        price_sum = state.pop("_price_sum")
        qty = state.pop("_qty")
        result.append(
            {
                **state,
                "quantity": qty,
                "price": round(price_sum / count, 4) if count > 0 else None,
            }
        )
    return result
